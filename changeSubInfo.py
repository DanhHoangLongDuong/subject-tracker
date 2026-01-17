from openpyxl import load_workbook

wb = load_workbook("C:/Users/Long/Desktop/subject-tracker/Subject/subject.xlsx")
ws = wb.active

Test = False
subjectName = int(input("Search subject by:\n1. Name\n2.Subject no.\nOption:"))
if subjectName == 1:
    subject = input("please type subject name: ")
    for i in range(2, ws.max_row + 1):
        subjectValue = ws.cell(row = i, column = 2).value
        if subjectValue == subject:
            subjectRow = i
            Test = True
elif subjectName == 2:
    subjectNo = input("please type subject no.: ")
    for i in range (2, ws.max_row + 1):
        subjectValue = ws.cell(row = i, column = 3).value
        if subjectValue == subjectNo:
            subjectRow = i
            Test = True

if Test == True:
    mark = int(input("New mark:"))
    ws.cell(row=subjectRow, column=4).value = mark
    if mark <= 100 and mark >= 85:
        ws.cell(row=subjectRow, column=5).value = "High Distinction (HD)"
    elif mark <= 84 and mark >= 75:
        ws.cell(row=subjectRow, column=5).value = "Distinction (D)"
    elif mark <= 74 and mark >= 65:
        ws.cell(row=subjectRow, column=5).value = "Credit (C)"
    elif mark <= 64 and mark >= 50:
        ws.cell(row=subjectRow, column=5).value = "Pass (P)"
    elif mark <= 49:
        ws.cell(row=subjectRow, column=5).value = "Fail (Z)"
elif Test == False:
    print("subject not found")

wb.save("C:/Users/Long/Desktop/subject-tracker/Subject/subject.xlsx")