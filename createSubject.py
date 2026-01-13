from openpyxl import load_workbook

wb = load_workbook("C:/Users/Long/Desktop/subject-tracker/Subject/subject.xlsx")
ws = wb.active

Row = ws.max_row

while True:
    Row += 1
    confirmation = int(input("1. Add subject information\n2. Out\nOption: "))
    if confirmation == 2:
        break
    elif confirmation == 1:
        ws.cell(row=Row, column=1).value = Row-1
        subjectName = input("Subject Name:")
        ws.cell(row=Row, column=2).value = subjectName
        subjectCode = input("subject no.: ")
        ws.cell(row=Row, column=3).value = subjectCode

        mark = int(input("mark: "))
        ws.cell(row=Row, column=4).value = mark

        if mark <= 100 and mark >= 85:
            ws.cell(row=Row, column=5).value = "High Distinction (HD)"
        elif mark <= 84 and mark >= 75:
            ws.cell(row=Row, column=5).value = "Distinction (D)"
        elif mark <= 74 and mark >= 65:
            ws.cell(row=Row, column=5).value = "Credit (C)"
        elif mark <= 64 and mark >= 50:
            ws.cell(row=Row, column=5).value = "Pass (P)"
        elif mark <= 49:
            ws.cell(row=Row, column=5).value = "Fail (Z)"
        ws.cell(row=Row, column=6).value = "6.0"
    


wb.save("C:/Users/Long/Desktop/subject-tracker/Subject/subject.xlsx")