from openpyxl import load_workbook

wb = load_workbook("C:/Users/Long/Desktop/subject-tracker/Subject/subject.xlsx")
ws = wb.active

overallMark = 0
for i in range (2, ws.max_row + 1):
    overallMark = overallMark + int(ws.cell(row = i, column=4).value)


    if overallMark <= 100 and overallMark >= 85:
        grade = "High Distinction (HD)"
    elif overallMark <= 84 and overallMark >= 75:
        grade = "Distinction (D)"
    elif overallMark <= 74 and overallMark >= 65:
        grade = "Credit (C)"
    elif overallMark <= 64 and overallMark >= 50:
        grade = "Pass (P)"
    elif overallMark <= 49:
        grade = "Fail (Z)"

Credit = (ws.max_row - 1)  * 6

print(f"Overall grade: {grade}\nCredit: {Credit}")